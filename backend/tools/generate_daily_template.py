"""
generate_daily_template.py
สร้างไฟล์ Excel template แอดมินใช้ลงข้อมูลรายวัน · เป็น backup ถ้าระบบหลักล่ม

ออกแบบให้เหมือนไฟล์ที่ admin ใช้อยู่ปัจจุบัน · 5 sheets:
  1. สต็อกหลัก       · ยอดทั้งหมดบริษัท + คลังกลาง (formula link) + สินค้ารอตรวจสอบ
  2. เติมตู้          · คลังกลาง + ยอดเติมทุกตู้ + เติมแต่ละตู้ + ยอดคงเหลือ
  3. ช่องเติมตู้      · slot grid ต่อตู้ (วาดเปล่า · admin เขียนแผนผังเอง)
  4. ยอดการซื้อบริษัท · purchase log + ราคาขาย (มี formula ยอดตั้งต้น/ราคาทุน-ซอง)
  5. สต็อกย่อย       · placeholder ว่างเปล่า

Usage:
    cd backend/tools
    py generate_daily_template.py

Env vars (อ่านจาก deploy/.env.local อัตโนมัติ):
    SUPABASE_URL (หรือ NEXT_PUBLIC_SUPABASE_URL)
    SUPABASE_SERVICE_KEY (หรือ NEXT_PUBLIC_SUPABASE_ANON_KEY · read-only พอ)

Output: DivisionX_Daily_Log_YYYY-MM-DD.xlsx
"""

import os
import sys
import io
from datetime import datetime
from pathlib import Path

# กัน UnicodeEncodeError บน Windows console (cp1252)
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

try:
    from dotenv import load_dotenv
    for env_path in [
        Path(__file__).parent / ".env",
        Path(__file__).parent.parent.parent / "deploy" / "scraper" / ".env",
        Path(__file__).parent.parent.parent / "deploy" / ".env.local",
    ]:
        if env_path.exists():
            load_dotenv(env_path)
            break
except ImportError:
    pass

from supabase import create_client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── Family mapping · จัดกลุ่ม SKU ตาม family เหมือนไฟล์ admin ───────
# ลำดับ family + วิธี match SKU + วิธี sort ภายใน family
SERIES_PRIORITY = {"OP": 1, "PRB": 2, "EB": 3, "FB": 1}  # ลำดับภายใน family

def _sort_key(s):
    """sort within family: series priority แล้วถึง sku_id"""
    series = s.get("series") or ""
    prio = SERIES_PRIORITY.get(series, 9)
    return (prio, s["sku_id"])

FAMILIES = [
    {
        "name": "One Piece",
        "match": lambda s: s["series"] in ("OP", "PRB", "EB"),
        "sort_key": _sort_key,
    },
    {
        "name": "Dragon Ball",
        "match": lambda s: s["series"] == "FB" or s["sku_id"].strip().upper() == "B29",
        # FB ก่อน B29
        "sort_key": lambda s: (0 if s["series"] == "FB" else 1, s["sku_id"]),
    },
    {
        "name": "Naruto",
        "match": lambda s: s["sku_id"].startswith("NRT"),
        # Jin ก่อน Series
        "sort_key": lambda s: (0 if "Jin" in s["sku_id"] else 1, s["sku_id"]),
    },
    {
        "name": "Pokemon",
        "match": lambda s: s["sku_id"].startswith("PKM"),
        "sort_key": lambda s: s["sku_id"],
    },
    {
        "name": "Solo Leveling",
        "match": lambda s: s["sku_id"].startswith("SLL"),
        "sort_key": lambda s: s["sku_id"],
    },
]


def short_sku(sku_id: str) -> str:
    """แปลง 'OP 01' → 'OP01', 'NRT Series - 02' → 'Series02', etc · ให้สั้นเหมือนไฟล์ admin"""
    s = sku_id.strip()
    if s.startswith("NRT "):
        rest = s[4:].replace("- ", "").replace(" ", "")  # 'Series - 01' → 'Series01'
        return rest
    if s.startswith("PKM "):
        rest = s[4:]
        if "Dream" in rest: return "Mega Dream"
        return rest  # "Ninja"
    if s.startswith("SLL "):
        return "Ua 1"  # ตามที่ admin ใช้
    # OP / PRB / EB / FB / B29 — เอา space ออก
    return s.replace(" ", "")


# ── Styles ─────────────────────────────────────────────────────────
TH_FONT = "Tahoma"
HEADER_FONT = Font(name=TH_FONT, size=11, bold=True, color="FFFFFF")
TITLE_FONT  = Font(name=TH_FONT, size=14, bold=True)
SUB_HEADER_FONT = Font(name=TH_FONT, size=11, bold=True)
DEFAULT_FONT = Font(name=TH_FONT, size=10)
FAMILY_FONT = Font(name=TH_FONT, size=11, bold=True, color="0F172A")

COLOR = {
    "main":     "1E40AF",  # navy
    "refill":   "059669",  # green
    "slot":     "7C3AED",  # purple
    "purchase": "DC2626",  # red
    "sub":      "F1F5F9",  # light gray for sub-section
    "family":   "FEF3C7",  # cream for family rows
    "formula":  "F8FAFC",  # very light for auto-calc
}

THIN_BORDER = Border(
    left=Side(border_style="thin", color="D1D5DB"),
    right=Side(border_style="thin", color="D1D5DB"),
    top=Side(border_style="thin", color="D1D5DB"),
    bottom=Side(border_style="thin", color="D1D5DB"),
)


def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def style_header_cell(cell, color):
    cell.font = HEADER_FONT
    cell.fill = fill(color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def style_sub_header_cell(cell, color="sub"):
    cell.font = SUB_HEADER_FONT
    cell.fill = fill(COLOR[color] if color in COLOR else color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def style_data_cell(cell, formula=False):
    cell.font = DEFAULT_FONT
    cell.border = THIN_BORDER
    if formula:
        cell.fill = fill(COLOR["formula"])


def style_family_row(ws, row, num_cols, family_name):
    """แถวหัว family · เช่น 'One Piece', 'Dragon Ball'"""
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(COLOR["family"])
        cell.border = THIN_BORDER
    cell0 = ws.cell(row=row, column=1, value=family_name)
    cell0.font = FAMILY_FONT
    cell0.alignment = Alignment(horizontal="left", vertical="center", indent=1)


# ── Fetch reference data ───────────────────────────────────────────
def fetch_data():
    url = (os.environ.get("SUPABASE_URL")
           or os.environ.get("NEXT_PUBLIC_SUPABASE_URL"))
    key = (os.environ.get("SUPABASE_SERVICE_KEY")
           or os.environ.get("SUPABASE_KEY")
           or os.environ.get("NEXT_PUBLIC_SUPABASE_ANON_KEY"))
    if not url or not key:
        sys.exit(
            "❌ ต้องตั้ง env vars:\n"
            "   • SUPABASE_URL (หรือ NEXT_PUBLIC_SUPABASE_URL)\n"
            "   • SUPABASE_SERVICE_KEY (หรือ NEXT_PUBLIC_SUPABASE_ANON_KEY)"
        )
    sb = create_client(url, key)
    skus = (sb.table("skus")
            .select("sku_id, name, series, sell_price, cost_price, avg_cost, packs_per_box, boxes_per_cotton")
            .eq("is_active", True)
            .order("sku_id")
            .execute().data) or []
    machines = (sb.table("machines")
                .select("machine_id, name, location, brand, status")
                .eq("status", "active")
                .order("machine_id")
                .execute().data) or []
    return skus, machines


def group_skus_by_family(skus):
    """คืน [(family_name, [sku_dict, ...]), ...] เรียงตาม FAMILIES + ภายในเรียงตาม sku_id"""
    groups = []
    used = set()
    for fam in FAMILIES:
        members = [s for s in skus if s["sku_id"] not in used and fam["match"](s)]
        members.sort(key=fam.get("sort_key", lambda s: s["sku_id"]))
        for s in members:
            used.add(s["sku_id"])
        if members:
            groups.append((fam["name"], members))
    # SKU ที่ไม่ match → กลุ่ม "อื่นๆ"
    leftovers = [s for s in skus if s["sku_id"] not in used]
    if leftovers:
        groups.append(("อื่นๆ", leftovers))
    return groups


# ── Sheet: สต็อกหลัก ──────────────────────────────────────────────
def build_main_stock_sheet(wb, groups, today):
    """
    Layout (cols A-N):
      A: SKU / family header
      B-D: ยอดทั้งหมดของบริษัท (COTTON / BOX / Pack) · admin กรอก
      E-G: คลังกลาง (COTTON / BOX / Pack) · formula =เติมตู้!M/N/O column
      H-N: สินค้ารอการตรวจสอบ (วันที่/เรื่อง/สถานที่/รหัสสินค้า/HOLD/ชำรุด/สูญหาย)
    """
    ws = wb.create_sheet("สต็อกหลัก")

    # Row 1 · top group headers
    ws.cell(row=1, column=1, value="วันที่").font = SUB_HEADER_FONT
    ws.cell(row=1, column=1).border = THIN_BORDER
    style_header_cell(ws.cell(row=1, column=2, value="ยอดทั้งหมดของบริษัท"), COLOR["main"])
    ws.merge_cells("B1:D1")
    style_header_cell(ws.cell(row=1, column=5, value="คลังกลาง"), COLOR["refill"])
    ws.merge_cells("E1:G1")
    style_header_cell(ws.cell(row=1, column=8, value="สินค้ารอการตรวจสอบ"), COLOR["purchase"])
    ws.merge_cells("H1:N1")

    # Row 2 · today date + family label for each section
    ws.cell(row=2, column=1, value=today).number_format = "yyyy-mm-dd"
    ws.cell(row=2, column=1).font = DEFAULT_FONT
    ws.cell(row=2, column=1).border = THIN_BORDER
    for col in (2, 5):
        c = ws.cell(row=2, column=col, value="(สินค้าทุกประเภท)")
        c.font = DEFAULT_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    # Sub-headers ของช่อง "สินค้ารอตรวจสอบ"
    for i, h in enumerate(["วันที่", "เรื่อง", "สถานที่", "รหัสสินค้า", "HOLD", "ชำรุด", "สูญหาย"], start=8):
        style_sub_header_cell(ws.cell(row=2, column=i, value=h))

    # Row 3 · unit headers (COTTON / BOX / Pack)
    ws.cell(row=3, column=1).border = THIN_BORDER
    for col, label in [(2, "COTTON"), (3, "BOX"), (4, "Pack"),
                       (5, "COTTON"), (6, "BOX"), (7, "Pack")]:
        c = ws.cell(row=3, column=col, value=label)
        c.font = SUB_HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
        c.fill = fill(COLOR["sub"])

    # SKU rows · group by family · sub-header row + members
    r = 4
    family_starts = {}  # family_name → starting row · เพื่อ formula reference จาก refill sheet
    refill_row_map = {}  # sku_id → row ใน refill sheet (จะตั้งทีหลัง)

    for family_name, members in groups:
        family_starts[family_name] = r
        # Family sub-header row
        style_family_row(ws, r, 7, family_name)
        # column B-D ตามไฟล์ admin = family label ซ้ำ
        ws.cell(row=r, column=2, value=family_name).font = FAMILY_FONT
        ws.cell(row=r, column=5, value=family_name).font = FAMILY_FONT
        r += 1
        for sku in members:
            ws.cell(row=r, column=1, value=short_sku(sku["sku_id"])).font = DEFAULT_FONT
            ws.cell(row=r, column=1).border = THIN_BORDER
            # B-D: admin input
            for col in range(2, 5):
                style_data_cell(ws.cell(row=r, column=col))
            # E-G: formula link from เติมตู้ (จะเซ็ตทีหลังเมื่อรู้ row ใน refill)
            for col in range(5, 8):
                style_data_cell(ws.cell(row=r, column=col), formula=True)
            refill_row_map[sku["sku_id"]] = r  # ใช้ row เดียวกันทั้ง 2 sheet
            r += 1

    # คอลัมน์ H-N (สินค้ารอตรวจสอบ) — empty rows 30 + row "รวม"
    for hold_r in range(3, 33):
        for col in range(8, 15):
            style_data_cell(ws.cell(row=hold_r, column=col))
    # Summary row
    sum_row = 33
    ws.cell(row=sum_row, column=10, value="รวม").font = SUB_HEADER_FONT
    ws.cell(row=sum_row, column=11, value=f"=SUM(L3:L{sum_row - 1})")
    ws.cell(row=sum_row, column=12, value=f"=SUM(M3:M{sum_row - 1})")
    ws.cell(row=sum_row, column=13, value=f"=SUM(N3:N{sum_row - 1})")
    for col in range(10, 14):
        ws.cell(row=sum_row, column=col).fill = fill(COLOR["formula"])
        ws.cell(row=sum_row, column=col).border = THIN_BORDER

    # Column widths
    widths = [14, 10, 10, 10, 10, 10, 10, 14, 14, 14, 14, 8, 8, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "B4"
    return refill_row_map  # ส่งกลับเพื่อ refill_sheet ใช้ alignment


# ── Sheet: เติมตู้ ────────────────────────────────────────────────
def build_refill_sheet(wb, groups, machines, today, refill_row_map):
    """
    Layout columns:
      A: วันที่ / SKU
      B-D: คลังกลาง (COTTON / BOX / Pack) · admin กรอก
      E-F: ยอดเติมทุกตู้ (BOX / Pack) · formula sum
      G-* : เติมแต่ละตู้ (BOX / Pack) คู่ละ machine
      next 3: ยอดคงเหลือคลังกลาง (COTTON / BOX / Pack) · formula
    """
    ws = wb.create_sheet("เติมตู้")
    n_machines = len(machines)

    # Column layout
    col_warehouse = 2          # B-D: COTTON BOX Pack
    col_total = 5              # E-F: ยอดเติมทุกตู้ BOX/Pack
    col_machines_start = 7     # G onwards: per machine BOX/Pack
    col_remain = col_machines_start + n_machines * 2  # COTTON BOX Pack

    # Row 1 · top group headers
    ws.cell(row=1, column=1, value="วันที่").font = SUB_HEADER_FONT
    ws.cell(row=1, column=1).border = THIN_BORDER
    style_header_cell(ws.cell(row=1, column=col_warehouse, value="คลังกลาง"), COLOR["refill"])
    ws.merge_cells(start_row=1, start_column=col_warehouse, end_row=1, end_column=col_warehouse + 2)
    style_header_cell(ws.cell(row=1, column=col_total, value="ยอดเติมทุกตู้"), COLOR["main"])
    ws.merge_cells(start_row=1, start_column=col_total, end_row=1, end_column=col_total + 1)

    for i, m in enumerate(machines):
        start = col_machines_start + i * 2
        title = f"เติม {m.get('name') or m['machine_id']}"
        style_header_cell(ws.cell(row=1, column=start, value=title), COLOR["slot"])
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=start + 1)

    style_header_cell(ws.cell(row=1, column=col_remain, value="ยอดคงเหลือคลังกลาง"), COLOR["refill"])
    ws.merge_cells(start_row=1, start_column=col_remain, end_row=1, end_column=col_remain + 2)

    # Row 2 · date row + family labels
    ws.cell(row=2, column=1, value=today).number_format = "yyyy-mm-dd"
    ws.cell(row=2, column=1).border = THIN_BORDER

    # Row 3 · unit headers
    units_warehouse = ["COTTON", "BOX", "Pack"]
    for i, u in enumerate(units_warehouse):
        c = ws.cell(row=3, column=col_warehouse + i, value=u)
        style_sub_header_cell(c)
    for i, u in enumerate(["BOX", "Pack"]):
        c = ws.cell(row=3, column=col_total + i, value=u)
        style_sub_header_cell(c)
    for i in range(n_machines):
        start = col_machines_start + i * 2
        style_sub_header_cell(ws.cell(row=3, column=start, value="BOX"))
        style_sub_header_cell(ws.cell(row=3, column=start + 1, value="Pack"))
    for i, u in enumerate(units_warehouse):
        c = ws.cell(row=3, column=col_remain + i, value=u)
        style_sub_header_cell(c)

    total_cols = col_remain + 2

    # SKU rows · same row numbers as สต็อกหลัก (formula link by row)
    r = 4
    for family_name, members in groups:
        style_family_row(ws, r, total_cols, family_name)
        ws.cell(row=r, column=col_warehouse, value=family_name).font = FAMILY_FONT
        ws.cell(row=r, column=col_total, value=family_name).font = FAMILY_FONT
        for i, m in enumerate(machines):
            ws.cell(row=r, column=col_machines_start + i * 2, value=family_name).font = FAMILY_FONT
        ws.cell(row=r, column=col_remain, value=family_name).font = FAMILY_FONT
        r += 1
        for sku in members:
            ws.cell(row=r, column=1, value=short_sku(sku["sku_id"])).font = DEFAULT_FONT
            ws.cell(row=r, column=1).border = THIN_BORDER

            # B-D: admin input คลังกลาง
            for col in range(col_warehouse, col_warehouse + 3):
                style_data_cell(ws.cell(row=r, column=col))

            # E-F: ยอดเติมทุกตู้ = sum BOX / Pack across machines
            box_refs = []
            pack_refs = []
            for i in range(n_machines):
                bx = get_column_letter(col_machines_start + i * 2)
                pk = get_column_letter(col_machines_start + i * 2 + 1)
                box_refs.append(f"{bx}{r}")
                pack_refs.append(f"{pk}{r}")
            ws.cell(row=r, column=col_total, value=f"={'+'.join(box_refs)}")
            ws.cell(row=r, column=col_total + 1, value=f"={'+'.join(pack_refs)}")
            for col in (col_total, col_total + 1):
                style_data_cell(ws.cell(row=r, column=col), formula=True)

            # Per-machine BOX/Pack · admin input
            for i in range(n_machines):
                for j in range(2):
                    style_data_cell(ws.cell(row=r, column=col_machines_start + i * 2 + j))

            # Remaining (COTTON/BOX/Pack) = warehouse - withdrawn
            col_b = get_column_letter(col_warehouse)
            col_c = get_column_letter(col_warehouse + 1)
            col_d = get_column_letter(col_warehouse + 2)
            col_e = get_column_letter(col_total)
            col_f = get_column_letter(col_total + 1)
            ws.cell(row=r, column=col_remain, value=f"={col_b}{r}")          # COTTON ยกมา
            ws.cell(row=r, column=col_remain + 1, value=f"={col_c}{r}-{col_e}{r}")  # BOX - withdrawn
            ws.cell(row=r, column=col_remain + 2, value=f"={col_d}{r}-{col_f}{r}")  # Pack - withdrawn
            for col in range(col_remain, col_remain + 3):
                style_data_cell(ws.cell(row=r, column=col), formula=True)

            # NOTE: row r ใน sheet นี้ตรงกับ row ใน สต็อกหลัก ที่เก็บไว้ใน refill_row_map (เพราะใช้ลำดับเดียวกัน)
            r += 1

    # อัพเดท formula ใน สต็อกหลัก ให้ point มาที่ row ของ refill (เท่ากัน)
    main_ws = wb["สต็อกหลัก"]
    for sku_id, main_row in refill_row_map.items():
        # E (COTTON) = 'เติมตู้'!M{r}, F (BOX) = N{r}, G (Pack) = O{r}
        col_m = get_column_letter(col_remain)
        col_n = get_column_letter(col_remain + 1)
        col_o = get_column_letter(col_remain + 2)
        main_ws.cell(row=main_row, column=5, value=f"='เติมตู้'!{col_m}{main_row}")
        main_ws.cell(row=main_row, column=6, value=f"='เติมตู้'!{col_n}{main_row}")
        main_ws.cell(row=main_row, column=7, value=f"='เติมตู้'!{col_o}{main_row}")

    # Column widths
    ws.column_dimensions["A"].width = 14
    for col in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.freeze_panes = "B4"
    return ws


# ── Sheet: ช่องเติมตู้ ────────────────────────────────────────────
def build_slot_grid_sheet(wb, machines, current_slots_by_machine):
    """
    Layout (สวยงาม + แยก brand):

      Row 1-2: Title + legend
      [brand banner VMS]
        [machine card 1] · slots organized as BOX section + PACK section
        [machine card 2]
        ...
      [brand banner WW]
        [machine card 5]
        ...

    Col A = row labels (ช่อง / สินค้า / SKU / ความจุ) · ทำให้อ่านง่าย
    Col B onwards = slot grid
    """
    ws = wb.create_sheet("ช่องเติมตู้")
    SLOTS_PER_ROW = 10  # ต่อบล็อก
    TOTAL_COLS = SLOTS_PER_ROW + 1  # +1 สำหรับ row label

    # Brand colors
    BRAND_COLOR = {
        "vms":       "1E40AF",  # navy blue
        "worldwide": "EA580C",  # orange
    }
    BRAND_NAME = {
        "vms":       "VMS · Inboxcorp",
        "worldwide": "Worldwide (WW)",
    }
    BOX_BG = "DBEAFE"   # light blue
    PACK_BG = "FEF3C7"  # light cream

    # ── Title + legend ─────────────────────────────────────────────
    title_cell = ws.cell(row=1, column=1, value="แผนผังช่องเติมตู้ — ของแต่ละตู้")
    title_cell.font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    ws.row_dimensions[1].height = 26

    # Legend
    legend = ws.cell(row=2, column=1,
        value="🔵 BOX (กล่อง · cap=4) · 🟡 PACK (ซอง · cap=12) · row 1=ช่อง · row 2=สินค้า · row 3=SKU+ประเภท · row 4=ความจุ")
    legend.font = Font(name=TH_FONT, size=9, italic=True, color="64748B")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)

    # ── Group machines by brand ────────────────────────────────────
    by_brand = {"vms": [], "worldwide": []}
    for m in machines:
        b = (m.get("brand") or "vms").lower()
        if b not in by_brand: by_brand[b] = []
        by_brand[b].append(m)

    r = 4  # start row

    def write_brand_banner(row, brand):
        color = BRAND_COLOR.get(brand, "475569")
        name = BRAND_NAME.get(brand, brand.upper())
        count = len(by_brand.get(brand, []))
        cell = ws.cell(row=row, column=1, value=f"  ▌ {name}  ·  {count} ตู้")
        cell.font = Font(name=TH_FONT, size=13, bold=True, color="FFFFFF")
        cell.fill = fill(color)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
        ws.row_dimensions[row].height = 28

    def write_machine_card(row, m):
        """เขียน 1 ตู้ · คืน row ถัดไปที่ใช้ได้"""
        nonlocal_brand = (m.get("brand") or "vms").lower()
        accent = BRAND_COLOR.get(nonlocal_brand, "475569")

        # Machine title bar
        title = f"📍 {m.get('name') or m['machine_id']}  ·  {m.get('location') or ''}"
        sub = f"({m['machine_id']})"
        tc = ws.cell(row=row, column=1, value=title)
        tc.font = Font(name=TH_FONT, size=12, bold=True, color="FFFFFF")
        tc.fill = fill(accent)
        tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS - 2)
        sc = ws.cell(row=row, column=TOTAL_COLS - 1, value=sub)
        sc.font = Font(name=TH_FONT, size=10, color="FFFFFF", italic=True)
        sc.fill = fill(accent)
        sc.alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells(start_row=row, start_column=TOTAL_COLS - 1, end_row=row, end_column=TOTAL_COLS)
        ws.row_dimensions[row].height = 22
        row += 1

        # ดึง slot ของตู้นี้ · sort by slot_number
        slots = current_slots_by_machine.get(m["machine_id"], [])
        slots = sorted(slots, key=lambda s: (s.get("slot_number") or ""))

        # ถ้าไม่มี · สร้าง 60 ช่องเปล่า
        if not slots:
            slots = [{"slot_number": str(i + 1).zfill(3), "product_name": "", "sku_id": "", "max_capacity": ""} for i in range(60)]

        # แบ่ง BOX vs PACK (จากชื่อสินค้า · ถ้าว่างให้ลง PACK)
        box_slots = [s for s in slots if "box" in (s.get("product_name") or "").lower()]
        pack_slots = [s for s in slots if "box" not in (s.get("product_name") or "").lower()]

        # Helper · render slot blocks · 4-row blocks
        def render_section(section_name, section_slots, bg_color, start_row):
            if not section_slots:
                return start_row
            # Section sub-header
            label = ws.cell(row=start_row, column=1, value=f"  {section_name}")
            label.font = Font(name=TH_FONT, size=10, bold=True, color="1E293B")
            label.fill = fill(bg_color)
            label.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=TOTAL_COLS)
            ws.row_dimensions[start_row].height = 18
            start_row += 1

            for block_start in range(0, len(section_slots), SLOTS_PER_ROW):
                block = section_slots[block_start:block_start + SLOTS_PER_ROW]

                # row labels in col A
                labels = ["ช่อง", "สินค้า", "SKU", "ความจุ"]
                for i, lbl in enumerate(labels):
                    lc = ws.cell(row=start_row + i, column=1, value=lbl)
                    lc.font = Font(name=TH_FONT, size=9, bold=True, color="475569")
                    lc.alignment = Alignment(horizontal="right", vertical="center")
                    lc.fill = fill("F1F5F9")
                    lc.border = THIN_BORDER

                for col_i, slot in enumerate(block):
                    col = col_i + 2  # +2 because col 1 is label
                    slot_num = slot.get("slot_number") or ""
                    product_name = slot.get("product_name") or ""
                    sku_id = slot.get("sku_id") or ""
                    cap = slot.get("max_capacity") or ""
                    is_box = "box" in (product_name or "").lower()
                    short = short_sku(sku_id) if sku_id else ""
                    short_with_type = f"{short} {'BOX' if is_box else 'PACK'}" if short else ""

                    # row 0: slot#
                    cn = ws.cell(row=start_row, column=col, value=slot_num)
                    cn.font = Font(name=TH_FONT, size=11, bold=True, color="0F172A")
                    cn.alignment = Alignment(horizontal="center", vertical="center")
                    cn.fill = fill("E2E8F0")
                    cn.border = THIN_BORDER

                    # row 1: family · derive series from sku_id prefix
                    family = ""
                    if sku_id:
                        sid = sku_id.strip()
                        derived_series = ""
                        for prefix in ("OP", "PRB", "EB", "FB"):
                            if sid.startswith(prefix + " ") or sid.startswith(prefix):
                                derived_series = prefix
                                break
                        for fam in FAMILIES:
                            if fam["match"]({"sku_id": sid, "series": derived_series}):
                                family = fam["name"]
                                break
                    cf = ws.cell(row=start_row + 1, column=col, value=family)
                    cf.font = Font(name=TH_FONT, size=9, color="64748B")
                    cf.alignment = Alignment(horizontal="center", vertical="center")
                    cf.fill = fill(bg_color)
                    cf.border = THIN_BORDER

                    # row 2: SKU + type
                    cs = ws.cell(row=start_row + 2, column=col, value=short_with_type)
                    cs.font = Font(name=TH_FONT, size=10, bold=True, color="0F172A")
                    cs.alignment = Alignment(horizontal="center", vertical="center")
                    cs.fill = fill(bg_color)
                    cs.border = THIN_BORDER

                    # row 3: capacity
                    cc = ws.cell(row=start_row + 3, column=col, value=cap)
                    cc.font = Font(name=TH_FONT, size=10, bold=True, color="DC2626")
                    cc.alignment = Alignment(horizontal="center", vertical="center")
                    cc.fill = fill("FFFFFF")
                    cc.border = THIN_BORDER

                # Fill empty cols ถ้า block สั้นกว่า SLOTS_PER_ROW
                for col_i in range(len(block), SLOTS_PER_ROW):
                    col = col_i + 2
                    for off in range(4):
                        ws.cell(row=start_row + off, column=col).border = THIN_BORDER

                ws.row_dimensions[start_row].height = 20
                ws.row_dimensions[start_row + 1].height = 14
                ws.row_dimensions[start_row + 2].height = 18
                ws.row_dimensions[start_row + 3].height = 16
                start_row += 4

            return start_row

        row = render_section("🔵 BOX section", box_slots, BOX_BG, row)
        row = render_section("🟡 PACK section", pack_slots, PACK_BG, row)
        return row + 1  # spacer

    # ── Render VMS section ─────────────────────────────────────────
    if by_brand.get("vms"):
        write_brand_banner(r, "vms")
        r += 1
        for m in by_brand["vms"]:
            r = write_machine_card(r, m)

    # ── Render WW section ──────────────────────────────────────────
    if by_brand.get("worldwide"):
        r += 1  # extra spacer
        write_brand_banner(r, "worldwide")
        r += 1
        for m in by_brand["worldwide"]:
            r = write_machine_card(r, m)

    # ── Column widths ──────────────────────────────────────────────
    ws.column_dimensions["A"].width = 11  # row label col
    for col in range(2, TOTAL_COLS + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13

    ws.freeze_panes = "B4"
    return ws


# ── Sheet: ยอดการซื้อบริษัท ───────────────────────────────────────
def build_purchase_sheet(wb, skus):
    """
    Left section (A-J): ยอดเบิกบริษัท
      A: วันที่ · B: ลำดับ · C: เรื่อง (family) · D: รหัสสินค้า
      E: Cotton · F: BOX · G: PACK
      H: ยอดตั้งต้นบริษัท (formula: E*ppc OR F*ppb+G)
      I: ราคาทุน/ซอง (formula: J/H)
      J: ยอดลงทุน (admin input)
    Right section (N-V): ยอดราคาขาย
      N: ลำดับ · O: เรื่อง · P: รหัสสินค้า
      Q: Pack (= ppc) · R: ราคา (admin) · S: ราคาเฉลี่ย (R/Q)
      T: ราคาขาย +30% (S*1.3) · U: Pack price · V: Box price (=U*ppb)
    """
    ws = wb.create_sheet("ยอดการซื้อบริษัท")

    # Headers row 1
    style_header_cell(ws.cell(row=1, column=1, value="ยอดเบิกบริษัท"), COLOR["main"])
    ws.merge_cells("A1:J1")
    ws.cell(row=1, column=11, value="หมายเหตุ").font = SUB_HEADER_FONT
    ws.cell(row=1, column=14, value="วันที่").font = SUB_HEADER_FONT
    style_header_cell(ws.cell(row=1, column=14, value="ยอดราคาขาย (markup 30%)"), COLOR["purchase"])
    ws.merge_cells("N1:V1")

    # Sub-headers row 2
    left_headers = ["วันที่", "ลำดับ", "เรื่อง", "รหัสสินค้า", "Cotton", "BOX", "PACK",
                    "ยอดตั้งต้นบริษัท", "ราคาทุน/ซอง", "ยอดลงทุน"]
    for i, h in enumerate(left_headers, start=1):
        style_sub_header_cell(ws.cell(row=2, column=i, value=h))
    right_headers = ["ลำดับ", "เรื่อง", "รหัสสินค้า", "Pack", "ราคา",
                     "ราคาเฉลี่ย", "ยอดคงเหลือ +30%", "Pack", "Box"]
    for i, h in enumerate(right_headers, start=14):
        style_sub_header_cell(ws.cell(row=2, column=i, value=h))

    # Build SKU lookup for family + ppc
    def packs_per_cotton(s):
        return (s.get("packs_per_box") or 1) * (s.get("boxes_per_cotton") or 1)
    def family_of(s):
        for fam in FAMILIES:
            if fam["match"](s): return fam["name"]
        return "อื่นๆ"

    # SKU rows · order ตาม family แล้ว sku_id
    sorted_skus = []
    for fam_def in FAMILIES:
        members = [s for s in skus if fam_def["match"](s)]
        members.sort(key=fam_def.get("sort_key", lambda s: s["sku_id"]))
        sorted_skus.extend(members)
    leftover = [s for s in skus if s not in sorted_skus]
    sorted_skus.extend(leftover)

    r = 3
    for idx, s in enumerate(sorted_skus, start=1):
        ppc = packs_per_cotton(s)
        ppb = s.get("packs_per_box") or 1
        fam = family_of(s)
        short = short_sku(s["sku_id"])

        # Left section
        if idx == 1:
            ws.cell(row=r, column=1, value=datetime.now().date()).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2, value=idx)
        ws.cell(row=r, column=3, value=fam)
        ws.cell(row=r, column=4, value=short)
        # E/F/G admin input (Cotton/Box/Pack)
        for col in range(5, 8):
            style_data_cell(ws.cell(row=r, column=col))
        # H: formula ยอดตั้งต้น = E*ppc + F*ppb + G
        ws.cell(row=r, column=8, value=f"=E{r}*{ppc}+F{r}*{ppb}+G{r}")
        style_data_cell(ws.cell(row=r, column=8), formula=True)
        # I: formula ราคาทุน/ซอง = J/H
        ws.cell(row=r, column=9, value=f'=IF(H{r}>0,J{r}/H{r},"")')
        style_data_cell(ws.cell(row=r, column=9), formula=True)
        # J: admin input ยอดลงทุน
        style_data_cell(ws.cell(row=r, column=10))

        # Right section · ราคาขาย
        ws.cell(row=r, column=14, value=idx)
        ws.cell(row=r, column=15, value=fam)
        ws.cell(row=r, column=16, value=short)
        ws.cell(row=r, column=17, value=ppc)  # Q: Pack (= ppc)
        style_data_cell(ws.cell(row=r, column=18))  # R: ราคา (admin)
        # S = R/Q
        ws.cell(row=r, column=19, value=f'=IF(AND(Q{r}>0,R{r}<>""),R{r}/Q{r},"")')
        style_data_cell(ws.cell(row=r, column=19), formula=True)
        # T = S*1.3
        ws.cell(row=r, column=20, value=f'=IF(S{r}<>"",S{r}*1.3,"")')
        style_data_cell(ws.cell(row=r, column=20), formula=True)
        # U: Pack price admin / V: Box price = U*ppb
        ws.cell(row=r, column=21, value=s.get("sell_price"))
        style_data_cell(ws.cell(row=r, column=21))
        ws.cell(row=r, column=22, value=f'=IF(U{r}<>"",U{r}*{ppb},"")')
        style_data_cell(ws.cell(row=r, column=22), formula=True)

        r += 1

    # Summary row · sum J (ยอดลงทุน)
    ws.cell(row=r + 1, column=9, value="รวมยอดลงทุน").font = SUB_HEADER_FONT
    ws.cell(row=r + 1, column=10, value=f"=SUM(J3:J{r})")
    ws.cell(row=r + 1, column=10).font = SUB_HEADER_FONT
    ws.cell(row=r + 1, column=10).fill = fill(COLOR["formula"])

    # Column widths
    widths = [12, 6, 14, 12, 8, 8, 8, 16, 12, 14, 18, 4, 4, 6, 14, 12, 10, 10, 12, 14, 10, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    return ws


# ── Sheet: สต็อกย่อย (placeholder) ────────────────────────────────
def build_substock_sheet(wb):
    ws = wb.create_sheet("สต็อกย่อย")
    ws.cell(row=1, column=1, value="(พื้นที่ว่างสำหรับ admin บันทึก stock ย่อย)").font = DEFAULT_FONT
    ws.column_dimensions["A"].width = 60
    return ws


# ── Main ───────────────────────────────────────────────────────────
def main():
    print("📥 Fetching SKUs + Machines from Supabase...")
    skus, machines = fetch_data()
    print(f"   • {len(skus)} SKUs · {len(machines)} machines")

    # Optional · ดึง current slot mappings สำหรับ pre-fill ช่องเติมตู้
    print("📥 Fetching current slot mappings...")
    url = os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    key = (os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("SUPABASE_KEY")
           or os.environ.get("NEXT_PUBLIC_SUPABASE_ANON_KEY"))
    sb = create_client(url, key)
    try:
        ms = sb.table("machine_stock").select("machine_id, slot_number, product_name, sku_id, max_capacity").execute().data
        current_slots_by_machine = {}
        for row in ms:
            current_slots_by_machine.setdefault(row["machine_id"], []).append(row)
    except Exception as e:
        print(f"   ⚠ machine_stock fetch failed: {e} · ใช้ grid เปล่าแทน")
        current_slots_by_machine = {}

    print("📝 Building workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    groups = group_skus_by_family(skus)
    today = datetime.now().date()

    refill_row_map = build_main_stock_sheet(wb, groups, today)
    build_refill_sheet(wb, groups, machines, today, refill_row_map)
    build_slot_grid_sheet(wb, machines, current_slots_by_machine)
    build_purchase_sheet(wb, skus)
    build_substock_sheet(wb)

    date_str = today.strftime("%Y-%m-%d")
    out_path = Path(__file__).parent / f"DivisionX_Daily_Log_{date_str}.xlsx"
    wb.save(out_path)
    print(f"✅ Saved: {out_path}")
    print(f"   ขนาด: {out_path.stat().st_size:,} bytes")


if __name__ == "__main__":
    main()
