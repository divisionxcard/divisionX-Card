"""
Cutoff Re-Seed (V3) · Excel → Seed SQL with Cotton refill calculation

Logic:
  1) อ่าน Main_Stock จาก Excel → main_packs[sku] + cost[sku]
  2) อ่าน User_Stock จาก Excel → user_packs[(sku, username)]
  3) Generate SQL ที่ใน DB จะ:
       a) Query sales(2026-05-01..05) per SKU (ที่เกิดไปแล้ว)
       b) Query machine_stock per SKU (snapshot ตอน run)
       c) Compute total = main + user + machine + sales
       d) Compute target = ceil(total / packs_per_cotton) * packs_per_cotton
       e) Compute refill = target - total  (กรรมการเติมเพิ่มเท่านี้)
       f) RAISE NOTICE: รายงาน "ต้องเติมต่อ SKU" (paste แรก รัน·ดู·ROLLBACK)
       g) ถ้าเปลี่ยน mode = 'SEED' → INSERT seed (paste 2 หลังเติมจริง)

Usage:
  py 03_excel_to_sql_v3.py [path/to/golive_filled.xlsx]

Output:
  03a_refill_report.sql  — รัน "ก่อน" กรรมการเติม → ดูว่าต้องเติมเท่าไหร่
  03b_seed_v3.sql        — รัน "หลัง" กรรมการเติม → INSERT จริง (target qty)
"""

import sys
import io
from datetime import datetime
from openpyxl import load_workbook
from collections import defaultdict

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")


CUTOFF_TS  = "2026-05-06 00:00:00+07"
LOT_NUMBER = "CUTOFF-20260506"
SOURCE     = "Cutoff Re-Seed (2026-05-06)"
CREATED_BY = "system_cutoff"

SALES_FROM = "2026-05-01 00:00:00+07"
SALES_TO   = "2026-05-06 00:00:00+07"   # exclusive — ครอบคลุม 1, 2, 3, 4, 5

SKUS_INFO = {
    # (series, packs_per_box, boxes_per_cotton) → packs_per_cotton = ppb × bpc
    "OP 01":  ("OP",  24, 12), "OP 02":  ("OP",  24, 12), "OP 03":  ("OP",  24, 12),
    "OP 04":  ("OP",  24, 12), "OP 05":  ("OP",  24, 12), "OP 06":  ("OP",  24, 12),
    "OP 07":  ("OP",  24, 12), "OP 08":  ("OP",  24, 12), "OP 09":  ("OP",  24, 12),
    "OP 10":  ("OP",  24, 12), "OP 11":  ("OP",  24, 12), "OP 12":  ("OP",  24, 12),
    "OP 13":  ("OP",  24, 12), "OP 14":  ("OP",  24, 12), "OP 15":  ("OP",  24, 12),
    "EB 01":  ("EB",  24, 12), "EB 02":  ("EB",  24, 12), "EB 03":  ("EB",  24, 12),
    "EB 04":  ("EB",  24, 12),
    "PRB 01": ("PRB", 10, 10),  # 100 packs/cotton
    "PRB 02": ("PRB", 10, 20),  # 200 packs/cotton (ใหญ่กว่า PRB 01)
}
KNOWN_USERS = {"divisionxcard", "tueza5432", "aofwara66", "power23n", "mzadiz1989"}


def cell(ws, r, c):
    v = ws.cell(r, c).value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def parse_main(ws):
    main, cost, errs = {}, {}, []
    for r in range(4, 4 + 25):
        sku = cell(ws, r, 1)
        if not sku:
            continue
        if sku not in SKUS_INFO:
            errs.append(f"Main_Stock row {r}: sku_id '{sku}' ไม่รู้จัก")
            continue
        _, ppb, bpc = SKUS_INFO[sku]
        ppc = ppb * bpc
        cottons = cell(ws, r, 5) or 0
        boxes   = cell(ws, r, 6) or 0
        cost_pp = cell(ws, r, 8) or 0
        try:
            packs = int(cottons) * ppc + int(boxes) * ppb
        except (TypeError, ValueError):
            errs.append(f"Main_Stock row {r}: full_cottons/full_boxes ไม่ใช่ตัวเลข")
            continue
        if packs < 0:
            errs.append(f"Main_Stock row {r}: total_packs ติดลบ")
            continue
        main[sku] = packs
        try:
            cost[sku] = float(cost_pp) if cost_pp else 0.0
        except (TypeError, ValueError):
            errs.append(f"Main_Stock row {r}: unit_cost ไม่ใช่ตัวเลข")
    return main, cost, errs


def parse_user(ws):
    user_packs = defaultdict(int)
    errs = []
    consecutive_empty = 0
    for r in range(5, 200):
        username = cell(ws, r, 1)
        if not username:
            consecutive_empty += 1
            if consecutive_empty >= 3:
                break  # ติดต่อกันว่าง 3 row = จบ data block (กัน scan reference table)
            continue
        consecutive_empty = 0
        # หยุดก่อนถ้าเจอ header/reference table
        if username.lower() == "username" or "reference" in username.lower() or "👥" in username:
            break
        if username not in KNOWN_USERS:
            errs.append(f"User_Stock row {r}: username '{username}' ไม่รู้จัก")
            continue
        sku = cell(ws, r, 2)
        if not sku:
            errs.append(f"User_Stock row {r}: ไม่มี sku_id")
            continue
        if sku not in SKUS_INFO:
            errs.append(f"User_Stock row {r}: sku_id '{sku}' ไม่รู้จัก")
            continue
        _, ppb, bpc = SKUS_INFO[sku]
        ppc = ppb * bpc
        cottons = cell(ws, r, 3) or 0
        boxes   = cell(ws, r, 4) or 0
        loose   = cell(ws, r, 5) or 0
        try:
            packs = int(cottons) * ppc + int(boxes) * ppb + int(loose)
        except (TypeError, ValueError):
            errs.append(f"User_Stock row {r}: cottons/boxes/loose_packs ไม่ใช่ตัวเลข")
            continue
        if packs <= 0:
            continue
        user_packs[(sku, username)] += packs
    return dict(user_packs), errs


def s(v):
    if v is None:
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def build_excel_values(main, cost):
    """VALUES sub-query สำหรับใช้ใน SQL — main + ppc + cost ต่อ SKU"""
    rows = []
    for sku, info in SKUS_INFO.items():
        _, ppb, bpc = info
        ppc = ppb * bpc
        m = main.get(sku, 0)
        c = cost.get(sku, 0.0)
        rows.append(f"  ({s(sku)}, {ppc}, {m}, {c:.2f})")
    return ",\n".join(rows)


def build_user_values(user_packs):
    """VALUES sub-query สำหรับ stock_transfers"""
    if not user_packs:
        return None
    rows = []
    for (sku, username), packs in sorted(user_packs.items()):
        rows.append(f"  ({s(sku)}, {s(username)}, {packs})")
    return ",\n".join(rows)


def gen_refill_report(main, cost, user_packs):
    user_sum = defaultdict(int)
    for (sku, _), p in user_packs.items():
        user_sum[sku] += p

    out = []
    out.append("-- =============================================================")
    out.append("-- Cutoff Re-Seed · Step 3a: REFILL REPORT (run before refill)")
    out.append(f"-- Generated: {datetime.now().isoformat(timespec='seconds')}")
    out.append("-- =============================================================")
    out.append("-- รัน \"ก่อน\" กรรมการเติม · ดูว่า SKU ไหนต้องเติมกี่ packs ให้ครบ Cotton")
    out.append("-- ไม่ INSERT อะไร · แค่ SELECT")
    out.append("-- ⚠ ก่อน run · กดปุ่ม VMS sync ให้ machine_stock fresh")
    out.append("-- =============================================================")
    out.append("")
    out.append("WITH excel_data AS (")
    out.append("  SELECT * FROM (VALUES")
    out.append(build_excel_values(main, cost))
    out.append("  ) AS t(sku_id, packs_per_cotton, main_packs, unit_cost)")
    out.append("),")
    out.append("user_totals AS (")
    out.append("  SELECT sku_id, SUM(packs) AS user_packs FROM (VALUES")
    user_rows = []
    for sku in SKUS_INFO.keys():
        u = user_sum.get(sku, 0)
        user_rows.append(f"    ({s(sku)}, {u})")
    out.append(",\n".join(user_rows))
    out.append("  ) AS u(sku_id, packs) GROUP BY sku_id")
    out.append("),")
    out.append("sales_totals AS (")
    out.append("  SELECT sku_id, COALESCE(SUM(quantity_sold), 0) AS sold")
    out.append("  FROM sales")
    out.append(f"  WHERE sold_at >= '{SALES_FROM}'::timestamptz")
    out.append(f"    AND sold_at <  '{SALES_TO}'::timestamptz")
    out.append("    AND sku_id IS NOT NULL")
    out.append("  GROUP BY sku_id")
    out.append("),")
    out.append("machine_totals AS (")
    out.append("  -- Box slots (product_name มี 'box') → remain เป็นกล่อง · ต้อง × packs_per_box")
    out.append("  SELECT sku_id, SUM(CASE")
    out.append("    WHEN product_name ILIKE '%box%' AND sku_id LIKE 'PRB%' THEN remain * 10")
    out.append("    WHEN product_name ILIKE '%box%' THEN remain * 24")
    out.append("    ELSE remain")
    out.append("  END) AS machine_packs")
    out.append("  FROM machine_stock")
    out.append("  WHERE sku_id IS NOT NULL AND remain > 0")
    out.append("  GROUP BY sku_id")
    out.append(")")
    out.append("SELECT")
    out.append("  e.sku_id,")
    out.append("  e.packs_per_cotton AS ppc,")
    out.append("  e.main_packs       AS main_now,")
    out.append("  COALESCE(u.user_packs, 0)    AS user_now,")
    out.append("  COALESCE(m.machine_packs, 0) AS machine_now,")
    out.append("  COALESCE(s.sold, 0)          AS sold_1_5,")
    out.append("  e.main_packs + COALESCE(u.user_packs, 0) + COALESCE(m.machine_packs, 0) + COALESCE(s.sold, 0) AS total_now,")
    out.append("  CEIL((e.main_packs + COALESCE(u.user_packs, 0) + COALESCE(m.machine_packs, 0) + COALESCE(s.sold, 0))::numeric / e.packs_per_cotton) * e.packs_per_cotton AS target,")
    out.append("  CEIL((e.main_packs + COALESCE(u.user_packs, 0) + COALESCE(m.machine_packs, 0) + COALESCE(s.sold, 0))::numeric / e.packs_per_cotton) * e.packs_per_cotton")
    out.append("    - (e.main_packs + COALESCE(u.user_packs, 0) + COALESCE(m.machine_packs, 0) + COALESCE(s.sold, 0)) AS refill_packs,")
    out.append("  e.unit_cost")
    out.append("FROM excel_data e")
    out.append("LEFT JOIN user_totals u    ON u.sku_id = e.sku_id")
    out.append("LEFT JOIN sales_totals s   ON s.sku_id = e.sku_id")
    out.append("LEFT JOIN machine_totals m ON m.sku_id = e.sku_id")
    out.append("ORDER BY e.sku_id;")
    out.append("")
    return "\n".join(out)


def gen_seed(main, cost, user_packs):
    user_sum = defaultdict(int)
    for (sku, _), p in user_packs.items():
        user_sum[sku] += p

    out = []
    out.append("-- =============================================================")
    out.append("-- Cutoff Re-Seed · Step 3b: SEED (run AFTER refill)")
    out.append(f"-- Generated: {datetime.now().isoformat(timespec='seconds')}")
    out.append("-- =============================================================")
    out.append("-- รัน \"หลัง\" กรรมการเติมของจริงตามรายงาน 03a เสร็จแล้ว")
    out.append("-- ⚠ ก่อน run · ต้อง:")
    out.append("--   1) Reset แล้ว (02_reset_v3.sql)")
    out.append("--   2) Trigger VMS sync รอบสุดท้าย (machine_stock fresh)")
    out.append("--   3) แก้ Excel · update main_packs ให้รวมของที่เติมแล้ว")
    out.append("--   4) Re-run python converter ให้ Excel data ใหม่ลง SQL นี้")
    out.append("-- =============================================================")
    out.append("")
    out.append("BEGIN;")
    out.append("")

    # ── stock_in ──────────────────────────────────────────
    out.append("-- 1) stock_in: 1 row/SKU = main + user + machine + sales(1-5)")
    out.append("--    ใช้ unit_cost จาก Excel · total_cost = qty × cost")
    out.append("WITH excel_data AS (")
    out.append("  SELECT * FROM (VALUES")
    out.append(build_excel_values(main, cost))
    out.append("  ) AS t(sku_id, packs_per_cotton, main_packs, unit_cost)")
    out.append("),")
    out.append("user_totals AS (")
    out.append("  SELECT sku_id, SUM(packs) AS user_packs FROM (VALUES")
    user_rows = []
    for sku in SKUS_INFO.keys():
        u = user_sum.get(sku, 0)
        user_rows.append(f"    ({s(sku)}, {u})")
    out.append(",\n".join(user_rows))
    out.append("  ) AS u(sku_id, packs) GROUP BY sku_id")
    out.append("),")
    out.append("sales_totals AS (")
    out.append("  SELECT sku_id, COALESCE(SUM(quantity_sold), 0) AS sold")
    out.append("  FROM sales")
    out.append(f"  WHERE sold_at >= '{SALES_FROM}'::timestamptz")
    out.append(f"    AND sold_at <  '{SALES_TO}'::timestamptz")
    out.append("    AND sku_id IS NOT NULL")
    out.append("  GROUP BY sku_id")
    out.append("),")
    out.append("machine_totals AS (")
    out.append("  -- Box slots (product_name มี 'box') → remain เป็นกล่อง · ต้อง × packs_per_box")
    out.append("  SELECT sku_id, SUM(CASE")
    out.append("    WHEN product_name ILIKE '%box%' AND sku_id LIKE 'PRB%' THEN remain * 10")
    out.append("    WHEN product_name ILIKE '%box%' THEN remain * 24")
    out.append("    ELSE remain")
    out.append("  END) AS machine_packs")
    out.append("  FROM machine_stock")
    out.append("  WHERE sku_id IS NOT NULL AND remain > 0")
    out.append("  GROUP BY sku_id")
    out.append(")")
    out.append("INSERT INTO stock_in (sku_id, lot_number, source, unit, quantity, quantity_packs,")
    out.append("                       unit_cost, total_cost, purchased_at, note, created_by)")
    out.append("SELECT")
    out.append("  e.sku_id,")
    out.append(f"  {s(LOT_NUMBER)} AS lot_number,")
    out.append(f"  {s(SOURCE)} AS source,")
    out.append("  'pack' AS unit,")
    out.append("  (e.main_packs + COALESCE(u.user_packs, 0) + COALESCE(m.machine_packs, 0) + COALESCE(s.sold, 0)) AS quantity,")
    out.append("  (e.main_packs + COALESCE(u.user_packs, 0) + COALESCE(m.machine_packs, 0) + COALESCE(s.sold, 0)) AS quantity_packs,")
    out.append("  e.unit_cost,")
    out.append("  (e.main_packs + COALESCE(u.user_packs, 0) + COALESCE(m.machine_packs, 0) + COALESCE(s.sold, 0)) * e.unit_cost AS total_cost,")
    out.append(f"  {s(CUTOFF_TS)}::timestamptz AS purchased_at,")
    out.append("  format('Cutoff · Main=%s User=%s Machine=%s Sold=%s',")
    out.append("    e.main_packs, COALESCE(u.user_packs, 0), COALESCE(m.machine_packs, 0), COALESCE(s.sold, 0)) AS note,")
    out.append(f"  {s(CREATED_BY)} AS created_by")
    out.append("FROM excel_data e")
    out.append("LEFT JOIN user_totals u    ON u.sku_id = e.sku_id")
    out.append("LEFT JOIN sales_totals s   ON s.sku_id = e.sku_id")
    out.append("LEFT JOIN machine_totals m ON m.sku_id = e.sku_id")
    out.append("WHERE (e.main_packs + COALESCE(u.user_packs, 0) + COALESCE(m.machine_packs, 0) + COALESCE(s.sold, 0)) > 0;")
    out.append("")

    # ── stock_transfers ──────────────────────────────────
    out.append("-- 2) stock_transfers: User holdings (จาก Excel)")
    user_vals = build_user_values(user_packs)
    if user_vals:
        out.append("INSERT INTO stock_transfers (sku_id, lot_number, to_user_id, unit,")
        out.append("                              quantity, quantity_packs, transferred_at,")
        out.append("                              note, created_by)")
        out.append("SELECT v.sku_id,")
        out.append(f"       {s(LOT_NUMBER)} AS lot_number,")
        out.append("       p.id AS to_user_id,")
        out.append("       'pack' AS unit,")
        out.append("       v.packs AS quantity,")
        out.append("       v.packs AS quantity_packs,")
        out.append(f"       {s(CUTOFF_TS)}::timestamptz AS transferred_at,")
        out.append("       'Cutoff transfer (Re-Seed)' AS note,")
        out.append(f"       {s(CREATED_BY)} AS created_by")
        out.append("FROM (VALUES")
        out.append(user_vals)
        out.append(") AS v(sku_id, username, packs)")
        out.append("JOIN profiles p ON p.username = v.username;")
    else:
        out.append("-- (no user stock)")
    out.append("")

    # ── stock_out (machine load = machine_now + sales) ──
    out.append("-- 3) stock_out: เติมตู้ = machine_now + sales(1-5)")
    out.append("--    เพราะ balance_machine = stock_out - sales · ต้องการ balance = machine_now")
    out.append("--    1 row/(sku, machine) · sum machine_stock + sales จัดกลุ่มตาม machine_id")
    out.append("INSERT INTO stock_out (sku_id, machine_id, quantity_packs,")
    out.append("                        withdrawn_at, withdrawn_by_user_id, note, created_by)")
    out.append("SELECT t.sku_id, t.machine_id, t.qty,")
    out.append(f"       {s(CUTOFF_TS)}::timestamptz, NULL,")
    out.append("       'Initial machine load (Cutoff Re-Seed)',")
    out.append(f"       {s(CREATED_BY)}")
    out.append("FROM (")
    out.append("  -- Box slots → remain เป็นกล่อง · ต้อง × packs_per_box")
    out.append("  SELECT sku_id, machine_id, SUM(CASE")
    out.append("    WHEN product_name ILIKE '%box%' AND sku_id LIKE 'PRB%' THEN remain * 10")
    out.append("    WHEN product_name ILIKE '%box%' THEN remain * 24")
    out.append("    ELSE remain")
    out.append("  END) AS qty")
    out.append("  FROM machine_stock")
    out.append("  WHERE sku_id IS NOT NULL AND remain > 0")
    out.append("  GROUP BY sku_id, machine_id")
    out.append("  UNION ALL")
    out.append("  SELECT sku_id, machine_id, COALESCE(SUM(quantity_sold), 0)")
    out.append("  FROM sales")
    out.append(f"  WHERE sold_at >= '{SALES_FROM}'::timestamptz")
    out.append(f"    AND sold_at <  '{SALES_TO}'::timestamptz")
    out.append("    AND sku_id IS NOT NULL AND machine_id IS NOT NULL")
    out.append("  GROUP BY sku_id, machine_id")
    out.append(") t")
    out.append("WHERE t.qty > 0;")
    out.append("")

    # ── update cost_price + avg_cost ─────────────────────
    out.append("-- 4) อัปเดต skus.cost_price + avg_cost = unit_cost จาก Excel")
    out.append("--    cost_price ใช้ใน v_stock_balance/Dashboard · avg_cost ใช้ใน profit calc")
    rows = []
    for sku, c in sorted(cost.items()):
        if c > 0:
            rows.append(f"UPDATE skus SET cost_price = {c:.2f}, avg_cost = {c:.2f} WHERE sku_id = {s(sku)};")
    if rows:
        out.extend(rows)
    else:
        out.append("-- (no unit_cost data)")
    out.append("")

    # ── Verify ──────────────────────────────────────────
    out.append("-- 5) Verify: balance ทุก SKU ไม่ติดลบ")
    out.append("DO $$")
    out.append("DECLARE v_neg INTEGER;")
    out.append("BEGIN")
    out.append("  SELECT COUNT(*) INTO v_neg FROM v_stock_balance WHERE balance < 0;")
    out.append("  IF v_neg > 0 THEN")
    out.append("    RAISE EXCEPTION 'Seed failed: % SKU มี balance ติดลบ', v_neg;")
    out.append("  END IF;")
    out.append("  RAISE NOTICE 'Seed OK';")
    out.append("END $$;")
    out.append("")
    out.append("COMMIT;")
    out.append("-- ROLLBACK;")
    out.append("")

    out.append("-- Summary จาก Excel:")
    out.append(f"--   Main packs:        {sum(main.values())}")
    out.append(f"--   User packs:        {sum(user_packs.values())}")
    out.append(f"--   stock_transfers:   {len(user_packs)} rows")
    out.append(f"--   SKUs ที่มี Main:    {len([s for s,p in main.items() if p > 0])}")

    return "\n".join(out)


def main():
    in_path = sys.argv[1] if len(sys.argv) > 1 else "cutoff_filled.xlsx"
    wb = load_workbook(in_path, data_only=False)
    main_packs, cost, e1 = parse_main(wb["Main_Stock"])
    user_packs, e2       = parse_user(wb["User_Stock"])

    errs = e1 + e2
    if errs:
        print("ERRORS — fix Excel ก่อน:", file=sys.stderr)
        for e in errs:
            print(f"  - {e}", file=sys.stderr)
        sys.exit(1)

    refill_sql = gen_refill_report(main_packs, cost, user_packs)
    seed_sql   = gen_seed(main_packs, cost, user_packs)

    with open("03a_refill_report.sql", "w", encoding="utf-8") as f:
        f.write(refill_sql)
    with open("03b_seed_v3.sql", "w", encoding="utf-8") as f:
        f.write(seed_sql)

    print("OK wrote 03a_refill_report.sql · 03b_seed_v3.sql")
    print(f"  - Main packs total:  {sum(main_packs.values())}")
    print(f"  - User packs total:  {sum(user_packs.values())}")
    print(f"  - User rows:         {len(user_packs)}")
    print(f"  - SKUs with cost:    {len([1 for c in cost.values() if c > 0])}/{len(cost)}")
    print()
    print("Next:")
    print("  1) Paste 03a_refill_report.sql ใน Supabase → ดูตัวเลข refill_packs")
    print("  2) กรรมการเติมของตามรายงาน → admin update Main_Stock ใน Excel")
    print("  3) Re-run script นี้ → 03b_seed_v3.sql รุ่นใหม่ (main_packs รวมที่เติมแล้ว)")
    print("  4) Paste 02_reset_v3.sql ใน Supabase → reset stock tables")
    print("  5) Paste 03b_seed_v3.sql ใน Supabase → seed final")


if __name__ == "__main__":
    main()
