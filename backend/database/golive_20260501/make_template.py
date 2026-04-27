"""
สร้าง Excel template สำหรับให้แอดมินกรอกข้อมูลตั้งต้นวัน Go-Live 1 พ.ค. 2026

2 sheets:
  1. Main_Stock — สต็อกในคลังหลัก (21 SKU pre-filled · ใส่ Cotton+Box+Pack+ราคา)
  2. User_Stock — สต็อกที่แอดมินแต่ละคนถือเหลือ (กรอกอิสระ)

(ตู้ไม่ต้องกรอก · ระบบดึงจาก VMS อัตโนมัติหลังกดปุ่ม sync)

Output: golive_template.xlsx (overwrite ถ้ามีอยู่แล้ว)
Run:    py make_template.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── ข้อมูลจาก prod (ณ 27 เม.ย. 2026) ─────────────────────────
SKUS = [
    # (sku_id, series, packs_per_box, boxes_per_cotton)
    ("OP 01",  "OP",  12, 12),
    ("OP 02",  "OP",  12, 12),
    ("OP 03",  "OP",  12, 12),
    ("OP 04",  "OP",  12, 12),
    ("OP 05",  "OP",  12, 12),
    ("OP 06",  "OP",  12, 12),
    ("OP 07",  "OP",  12, 12),
    ("OP 08",  "OP",  12, 12),
    ("OP 09",  "OP",  12, 12),
    ("OP 10",  "OP",  12, 12),
    ("OP 11",  "OP",  12, 12),
    ("OP 12",  "OP",  12, 12),
    ("OP 13",  "OP",  12, 12),
    ("OP 14",  "OP",  12, 12),
    ("OP 15",  "OP",  12, 12),
    ("EB 01",  "EB",  12, 12),
    ("EB 02",  "EB",  12, 12),
    ("EB 03",  "EB",  12, 12),
    ("EB 04",  "EB",  12, 12),
    ("PRB 01", "PRB", 10, 12),
    ("PRB 02", "PRB", 10, 12),
]

USERS = [
    # (username, display_name, role)
    ("divisionxcard",   "DivisionX Card",  "admin"),
    ("pornthep_sm1991", "T",               "admin"),
    ("aofwara66",       "AOFSEN",          "user"),
    ("power23n",        "พี่ M",          "user"),
    ("mzadiz1989",      "M-zadiz",         "user"),
]


# ── สีและสไตล์ ───────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="2C5282")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
INFO_FILL    = PatternFill("solid", fgColor="EDF2F7")  # column readonly
INPUT_FILL   = PatternFill("solid", fgColor="FFFAF0")  # column ที่ต้องกรอก
FORMULA_FILL = PatternFill("solid", fgColor="E6FFFA")  # auto-compute
NOTE_FONT    = Font(italic=True, color="718096", size=9)
THIN         = Side(style="thin", color="CBD5E0")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT         = Alignment(horizontal="left", vertical="center")


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER


def make_main_stock(wb):
    ws = wb.create_sheet("Main_Stock")
    # ⚠ Business rule: Main เก็บเฉพาะ Cotton + Box · ไม่มีซองเศษ
    # (เศษซอง = อยู่ที่ User · admin จ่ายของจาก Main เป็น Cotton/Box เต็มเท่านั้น)
    headers = [
        ("A", "sku_id",             INFO_FILL,    14),
        ("B", "series",             INFO_FILL,     9),
        ("C", "packs_per_box",      INFO_FILL,    13),
        ("D", "packs_per_cotton",   INFO_FILL,    16),
        ("E", "full_cottons",       INPUT_FILL,   13),
        ("F", "full_boxes",         INPUT_FILL,   12),
        ("G", "total_packs (auto)", FORMULA_FILL, 17),
        ("H", "unit_cost_per_pack", INPUT_FILL,   18),
        ("I", "note",               INPUT_FILL,   28),
    ]

    ws.merge_cells("A1:I1")
    ws["A1"] = "📦 สต็อกใน Main (คลังหลัก) ณ เช้าวันที่ 1 พ.ค. 2026"
    ws["A1"].font = Font(bold=True, size=12, color="2C5282")
    ws["A1"].alignment = LEFT

    ws.merge_cells("A2:I2")
    ws["A2"] = ("Main เก็บเฉพาะ Cotton + Box · ไม่มีซองเศษ (เศษอยู่ที่ User) · "
                "กรอก full_cottons (ลังใหญ่) + full_boxes (กล่อง) + unit_cost_per_pack (ราคาใบเสร็จล่าสุด)")
    ws["A2"].font = NOTE_FONT
    ws["A2"].alignment = LEFT

    HEADER_ROW = 3
    for col, label, _, width in headers:
        c = ws[f"{col}{HEADER_ROW}"]
        c.value = label
        style_header(c)
        ws.column_dimensions[col].width = width

    for i, (sku_id, series, ppb, bpc) in enumerate(SKUS):
        r = HEADER_ROW + 1 + i
        ws.cell(r, 1, sku_id).fill = INFO_FILL
        ws.cell(r, 2, series).fill = INFO_FILL
        ws.cell(r, 3, ppb).fill = INFO_FILL
        ws.cell(r, 4, ppb * bpc).fill = INFO_FILL  # packs_per_cotton
        ws.cell(r, 5, 0).fill = INPUT_FILL         # full_cottons
        ws.cell(r, 6, 0).fill = INPUT_FILL         # full_boxes
        ws.cell(r, 7, f"=E{r}*D{r}+F{r}*C{r}").fill = FORMULA_FILL  # total_packs
        ws.cell(r, 8, 0).fill = INPUT_FILL         # unit_cost_per_pack
        ws.cell(r, 9, "").fill = INPUT_FILL        # note
        for col_idx in range(1, 10):
            ws.cell(r, col_idx).border = BORDER
            ws.cell(r, col_idx).alignment = CENTER if col_idx <= 8 else LEFT

    ws.freeze_panes = f"A{HEADER_ROW + 1}"


def make_user_stock(wb):
    ws = wb.create_sheet("User_Stock")
    # User เก็บได้ทั้ง Cotton/Box/Pack · กรอกแยกหน่วย ระบบคำนวณรวมเป็นซอง
    headers = [
        ("A", "username",           INPUT_FILL,   18),
        ("B", "sku_id",             INPUT_FILL,   12),
        ("C", "full_cottons",       INPUT_FILL,   13),
        ("D", "full_boxes",         INPUT_FILL,   12),
        ("E", "loose_packs",        INPUT_FILL,   13),
        ("F", "total_packs (auto)", FORMULA_FILL, 17),
        ("G", "note",               INPUT_FILL,   28),
    ]

    ws.merge_cells("A1:G1")
    ws["A1"] = "👤 สต็อกที่แอดมินแต่ละคนถือเหลือ (หลังเติมตู้แล้ว) · ณ เช้า 1 พ.ค. 2026"
    ws["A1"].font = Font(bold=True, size=12, color="2C5282")

    ws.merge_cells("A2:G2")
    ws["A2"] = ("เพิ่ม 1 row ต่อ (ผู้ใช้, SKU) · กรอก full_cottons + full_boxes + loose_packs · "
                "ระบบคำนวณ total_packs ให้เอง (ดู SKU Reference ด้านล่างเพื่อเทียบขนาด)")
    ws["A2"].font = NOTE_FONT

    HEADER_ROW = 4
    for col, label, _, width in headers:
        c = ws[f"{col}{HEADER_ROW}"]
        c.value = label
        style_header(c)
        ws.column_dimensions[col].width = width

    # 60 empty input rows + auto formula สำหรับ total_packs
    for r in range(HEADER_ROW + 1, HEADER_ROW + 61):
        # ไม่ใส่ formula ที่นี่เพราะ packs_per_box ขึ้นกับ SKU · ใช้ VLOOKUP ไป SKU Reference
        # F = cottons * (ppb*12) + boxes * ppb + loose
        # ppb ดึงจาก SKU Reference ที่ row ของ ref_skus
        # ใช้ IFERROR เผื่อช่องว่าง
        ppb_lookup = f'IFERROR(VLOOKUP(B{r},$J$5:$K$25,2,FALSE),0)'
        ws.cell(r, 6, f'=IF(B{r}="","",C{r}*{ppb_lookup}*12+D{r}*{ppb_lookup}+E{r})').fill = FORMULA_FILL
        for col_idx in range(1, 8):
            if col_idx != 6:
                ws.cell(r, col_idx).fill = INPUT_FILL
            ws.cell(r, col_idx).border = BORDER

    # SKU Reference (ขวาของ data · column J-K) · ให้ VLOOKUP ใช้
    ws.cell(4, 10, "🎴 SKU Reference (อ้างอิง)").font = Font(bold=True, color="2C5282", size=10)
    ws.cell(4, 10).fill = INFO_FILL
    ws.merge_cells(start_row=4, start_column=10, end_row=4, end_column=11)
    for i, (sku_id, _, ppb, _) in enumerate(SKUS):
        r = 5 + i
        ws.cell(r, 10, sku_id).fill = INFO_FILL
        ws.cell(r, 11, ppb).fill = INFO_FILL
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 8

    # User reference table (ด้านล่าง)
    ref_start = HEADER_ROW + 63
    ws.cell(ref_start, 1, "👥 User Reference (อ้างอิง · อย่าแก้)").font = Font(bold=True, color="2C5282")
    ws.cell(ref_start + 1, 1, "username").fill = INFO_FILL
    ws.cell(ref_start + 1, 2, "display_name").fill = INFO_FILL
    ws.cell(ref_start + 1, 3, "role").fill = INFO_FILL
    for c in (1, 2, 3):
        ws.cell(ref_start + 1, c).font = Font(bold=True)
    for i, (u, dn, role) in enumerate(USERS):
        r = ref_start + 2 + i
        ws.cell(r, 1, u).fill = INFO_FILL
        ws.cell(r, 2, dn).fill = INFO_FILL
        ws.cell(r, 3, role).fill = INFO_FILL

    ws.freeze_panes = f"A{HEADER_ROW + 1}"


def make_readme(wb):
    ws = wb.create_sheet("README", 0)  # first tab
    ws.column_dimensions["A"].width = 115

    lines = [
        ("📋 DivisionX Card · Go-Live Template (1 พ.ค. 2026)", "title"),
        ("", None),
        ("Flow ใหม่ — เริ่มเติมตู้ก่อน · ระบบดึงสต็อกตู้จาก VMS อัตโนมัติ", "h2"),
        ("1) แอดมินเติมของจริงในตู้ทุกตู้ให้เต็ม (เช้า ~07:00-08:30)", None),
        ("2) ทีมเทคกด 'ดึงข้อมูล VMS' บนเว็บ → รอ ~1 นาที", None),
        ("3) ตรวจหน้า 'สต็อกหน้าตู้' ว่าระบบเห็นตรงกับ VMS", None),
        ("4) แอดมินนับสต็อกที่เหลือในมือ (User_Stock) + นับ Main → กรอกใน Excel นี้", None),
        ("5) Save → ส่งไฟล์ให้ทีมเทค → reset + seed → go-live", None),
        ("", None),
        ("วิธีกรอก Sheet 'Main_Stock'", "h2"),
        ("• Main เก็บเฉพาะ Cotton + Box · ไม่มีซองเศษ (เศษซองทั้งหมดอยู่ที่ User)", "note"),
        ("• full_cottons = ลังใหญ่ครบ (1 cotton = 12 box · OP/EB = 144 packs · PRB = 120 packs)", None),
        ("• full_boxes = กล่องครบที่ไม่ได้อยู่ในลัง", None),
        ("• total_packs = ระบบคำนวณให้อัตโนมัติ", None),
        ("• unit_cost_per_pack = ราคาทุน/ซอง (ใบเสร็จล่าสุด)", None),
        ("• ถ้าไม่มี SKU ไหนเลย ใส่ 0 ทุกช่อง", "note"),
        ("", None),
        ("วิธีกรอก Sheet 'User_Stock'", "h2"),
        ("• User เก็บได้ทั้ง Cotton/Box/Pack · กรอกแยกหน่วย", None),
        ("• username ต้องตรงกับ User Reference ด้านล่างของ sheet", None),
        ("• 1 row = 1 ผู้ใช้ × 1 SKU (เพิ่ม row ตามต้องการ)", None),
        ("• full_cottons + full_boxes + loose_packs = ของที่ผู้ใช้คนนั้น 'เหลือในมือ' หลังเติมตู้", None),
        ("• total_packs = ระบบคำนวณให้อัตโนมัติ (VLOOKUP packs_per_box จาก SKU Reference)", None),
        ("• ถ้าผู้ใช้ไม่มี SKU ใดเลย ข้ามได้", "note"),
        ("", None),
        ("วิธีกรอก Sheet 'User_Stock'", "h2"),
        ("• User เก็บได้ทั้ง Cotton/Box/Pack · นับรวมเป็น 'ซอง' ทั้งหมด", None),
        ("• ใส่จำนวน 'ซองรวม' ที่ผู้ใช้ถือเหลือในมือ (หลังเติมตู้แล้ว)", None),
        ("", None),
        ("ห้ามแก้", "h2"),
        ("• Sheet README นี้", None),
        ("• Column A-D ใน Main_Stock (sku_id / series / packs_per_box / packs_per_cotton)", None),
        ("• User Reference ด้านล่าง User_Stock", None),
        ("", None),
        ("คำถาม", "h2"),
        ("ถ้าไม่แน่ใจ → ทักเจ้าของระบบ (อย่าเดา) · ตัวเลขผิดยอดสต็อกพังหมดเดือน", None),
    ]

    for r, (text, kind) in enumerate(lines, start=1):
        c = ws.cell(r, 1, text)
        if kind == "title":
            c.font = Font(bold=True, size=16, color="1A365D")
        elif kind == "h2":
            c.font = Font(bold=True, size=12, color="2C5282")
        elif kind == "note":
            c.font = NOTE_FONT
        else:
            c.font = Font(size=11)
        c.alignment = LEFT


def main():
    wb = Workbook()
    wb.remove(wb.active)
    make_readme(wb)
    make_main_stock(wb)
    make_user_stock(wb)

    out = "golive_template.xlsx"
    wb.save(out)
    print(f"OK wrote {out}")


if __name__ == "__main__":
    main()
