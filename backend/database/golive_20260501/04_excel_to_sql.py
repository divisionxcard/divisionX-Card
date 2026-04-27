"""
Convert golive_template.xlsx (กรอกครบแล้ว) → seed SQL ready to paste

Logic:
  1) อ่าน Main_Stock     → main_packs[sku] + cost[sku]
  2) อ่าน User_Stock     → user_packs[(sku, user)]
  3) อ่าน Machine_Stock  → machine_packs[(sku, machine)] (รวม slot ที่ sku เดียวกันในตู้เดียว)
  4) total[sku] = Main + sum(User) + sum(Machine)
  5) Generate SQL:
       INSERT stock_in        : 1 row/SKU = total packs · unit_cost จาก Main
       INSERT stock_transfers : 1 row/(SKU, User) · ใช้ subquery หา user_id จาก username
       INSERT stock_out       : 1 row/(SKU, Machine) · withdrawn_by_user_id = NULL (Main → ตู้)
       INSERT machine_stock   : 1 row/slot (optional · ถ้ากรอก slot/sku/remain)
       UPDATE skus.avg_cost   : = unit_cost ของ Main (single lot ถือว่าตรง)

Output: 03_seed_generated.sql (overwrite)

Run: py 04_excel_to_sql.py [path/to/template.xlsx]
"""

import sys
from datetime import datetime
from openpyxl import load_workbook
from collections import defaultdict


GO_LIVE_TS  = "2026-05-01 08:00:00+07"
LOT_NUMBER  = "GOLIVE-20260501"
SOURCE      = "Initial Inventory (Go-Live 2026-05-01)"
CREATED_BY  = "system_golive"

# ── ข้อมูล reference (ต้องตรงกับ make_template.py) ─────────────
KNOWN_SKUS = {
    "OP 01", "OP 02", "OP 03", "OP 04", "OP 05", "OP 06", "OP 07",
    "OP 08", "OP 09", "OP 10", "OP 11", "OP 12", "OP 13", "OP 14", "OP 15",
    "EB 01", "EB 02", "EB 03", "EB 04",
    "PRB 01", "PRB 02",
}
KNOWN_USERS    = {"divisionxcard", "pornthep_sm1991", "aofwara66", "power23n", "mzadiz1989"}
KNOWN_MACHINES = {"chukes01", "chukes02", "chukes03", "chukes04"}


def cell(ws, row, col):
    """Get cell value · trim whitespace · return None if empty"""
    v = ws.cell(row, col).value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def parse_main(ws):
    """ใน Main_Stock: row 4 = header, row 5+ = data
       Columns: sku_id | series | packs_per_box | full_boxes | loose_packs | total_packs | unit_cost | note
    """
    main = {}        # sku → packs
    cost = {}        # sku → unit_cost
    errors = []
    for r in range(4, 4 + 25):  # ดูเผื่อ 25 rows
        sku = cell(ws, r, 1)
        if not sku:
            continue
        if sku not in KNOWN_SKUS:
            errors.append(f"Main_Stock row {r}: sku_id '{sku}' ไม่รู้จัก")
            continue
        ppb     = cell(ws, r, 3) or 0
        boxes   = cell(ws, r, 4) or 0
        loose   = cell(ws, r, 5) or 0
        cost_pp = cell(ws, r, 7) or 0

        # คำนวณ total_packs เอง (formula ใน Excel อาจไม่ evaluate ตอน load)
        try:
            total_packs = int(boxes) * int(ppb) + int(loose)
        except (TypeError, ValueError):
            errors.append(f"Main_Stock row {r}: full_boxes/loose_packs ไม่ใช่ตัวเลข")
            continue

        if total_packs < 0:
            errors.append(f"Main_Stock row {r}: total_packs ติดลบ")
            continue

        main[sku] = total_packs
        try:
            cost[sku] = float(cost_pp) if cost_pp else 0.0
        except (TypeError, ValueError):
            errors.append(f"Main_Stock row {r}: unit_cost_per_pack ไม่ใช่ตัวเลข")

    return main, cost, errors


def parse_user(ws):
    """ใน User_Stock: row 5 = header, row 6-55 = data
       Columns: username | display_name | sku_id | total_packs | note
    """
    user_packs = defaultdict(int)  # (sku, username) → packs
    errors = []
    for r in range(5, 56):
        username = cell(ws, r, 1)
        if not username:
            continue
        if username not in KNOWN_USERS:
            errors.append(f"User_Stock row {r}: username '{username}' ไม่รู้จัก")
            continue
        sku   = cell(ws, r, 3)
        packs = cell(ws, r, 4)
        if not sku:
            errors.append(f"User_Stock row {r}: ไม่มี sku_id")
            continue
        if sku not in KNOWN_SKUS:
            errors.append(f"User_Stock row {r}: sku_id '{sku}' ไม่รู้จัก")
            continue
        try:
            packs = int(packs or 0)
        except (TypeError, ValueError):
            errors.append(f"User_Stock row {r}: total_packs ไม่ใช่ตัวเลข")
            continue
        if packs <= 0:
            continue  # 0 = ข้าม
        user_packs[(sku, username)] += packs
    return dict(user_packs), errors


def parse_machine(ws):
    """ใน Machine_Stock: row 5 = header, row 6-105 = data
       Columns: machine_id | machine_name | slot_number | sku_id | remain | max_capacity | note
    """
    machine_packs = defaultdict(int)  # (sku, machine) → packs
    machine_slots = []                 # list ของ slot row สำหรับ insert machine_stock
    errors = []
    for r in range(5, 106):
        machine = cell(ws, r, 1)
        if not machine:
            continue
        if machine not in KNOWN_MACHINES:
            errors.append(f"Machine_Stock row {r}: machine_id '{machine}' ไม่รู้จัก")
            continue
        slot     = cell(ws, r, 3)
        sku      = cell(ws, r, 4)
        remain   = cell(ws, r, 5)
        max_cap  = cell(ws, r, 6)
        if not sku:
            continue  # slot ว่าง
        if sku not in KNOWN_SKUS:
            errors.append(f"Machine_Stock row {r}: sku_id '{sku}' ไม่รู้จัก")
            continue
        try:
            remain  = int(remain or 0)
            max_cap = int(max_cap or 0)
        except (TypeError, ValueError):
            errors.append(f"Machine_Stock row {r}: remain/max_capacity ไม่ใช่ตัวเลข")
            continue
        if remain < 0:
            errors.append(f"Machine_Stock row {r}: remain ติดลบ")
            continue
        if remain > 0:
            machine_packs[(sku, machine)] += remain
        machine_slots.append({
            "machine_id": machine,
            "slot_number": str(slot) if slot is not None else "",
            "sku_id": sku,
            "remain": remain,
            "max_capacity": max_cap,
        })
    return dict(machine_packs), machine_slots, errors


def sql_str(s):
    if s is None:
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


def gen_sql(main, cost, user_packs, machine_packs, machine_slots):
    out = []
    out.append("-- =============================================================")
    out.append("-- Go-Live 1 พ.ค. 2026 · Step 3: SEED (auto-generated)")
    out.append(f"-- Generated: {datetime.now().isoformat(timespec='seconds')}")
    out.append("-- =============================================================")
    out.append("-- Run หลัง 02_reset_transactional.sql · ใน prod SQL Editor")
    out.append("-- ตรวจ URL ว่าเป็น xethnqqmpvlpmafvphky · BEGIN ครอบทุก insert")
    out.append("-- =============================================================")
    out.append("")
    out.append("BEGIN;")
    out.append("")

    # ── 1) stock_in ──────────────────────────────────────────────
    out.append("-- 1) stock_in: Main + sum(User) + sum(Machine) per SKU")
    out.append("INSERT INTO stock_in (sku_id, source, unit, quantity, quantity_packs,")
    out.append("                       unit_cost, total_cost, purchased_at, note, created_by)")
    out.append("VALUES")
    rows = []
    user_sum    = defaultdict(int)
    machine_sum = defaultdict(int)
    for (sku, _), packs in user_packs.items():
        user_sum[sku] += packs
    for (sku, _), packs in machine_packs.items():
        machine_sum[sku] += packs
    all_skus = set(main) | set(user_sum) | set(machine_sum)
    for sku in sorted(all_skus):
        total = main.get(sku, 0) + user_sum.get(sku, 0) + machine_sum.get(sku, 0)
        if total <= 0:
            continue
        c = cost.get(sku, 0.0)
        rows.append(
            f"  ({sql_str(sku)}, {sql_str(SOURCE)}, 'pack', {total}, {total}, "
            f"{c:.2f}, {total * c:.2f}, {sql_str(GO_LIVE_TS)}, "
            f"{sql_str(f'Initial · Main={main.get(sku,0)} User={user_sum.get(sku,0)} Machine={machine_sum.get(sku,0)}')}, "
            f"{sql_str(CREATED_BY)})"
        )
    if not rows:
        out.append("  -- ⚠ ไม่มี stock_in (Excel ว่างเปล่า?)")
    else:
        out.append(",\n".join(rows) + ";")
    out.append("")

    # ── 2) stock_transfers (Main → User) ─────────────────────────
    out.append("-- 2) stock_transfers: ของที่ User แต่ละคนถือ (จาก Main)")
    if user_packs:
        out.append("INSERT INTO stock_transfers (sku_id, lot_number, to_user_id, unit,")
        out.append("                              quantity, quantity_packs, transferred_at,")
        out.append("                              note, created_by)")
        out.append("SELECT v.sku_id, v.lot_number, p.id, v.unit, v.quantity, v.quantity_packs,")
        out.append("       v.transferred_at, v.note, v.created_by")
        out.append("FROM (VALUES")
        rows = []
        for (sku, username), packs in sorted(user_packs.items()):
            rows.append(
                f"  ({sql_str(sku)}, {sql_str(LOT_NUMBER)}, {sql_str(username)}, "
                f"'pack', {packs}, {packs}, {sql_str(GO_LIVE_TS)}, "
                f"'Initial transfer (Go-Live)', {sql_str(CREATED_BY)})"
            )
        out.append(",\n".join(rows))
        out.append(") AS v(sku_id, lot_number, username, unit, quantity, quantity_packs,")
        out.append("       transferred_at, note, created_by)")
        out.append("JOIN profiles p ON p.username = v.username;")
    else:
        out.append("-- (no user stock)")
    out.append("")

    # ── 3) stock_out (Main → Machine pre-load) ───────────────────
    out.append("-- 3) stock_out: ของที่อยู่ในตู้แต่ละตู้ (Main → ตู้)")
    if machine_packs:
        out.append("INSERT INTO stock_out (sku_id, machine_id, quantity_packs,")
        out.append("                        withdrawn_at, withdrawn_by_user_id, note, created_by)")
        out.append("VALUES")
        rows = []
        for (sku, machine), packs in sorted(machine_packs.items()):
            rows.append(
                f"  ({sql_str(sku)}, {sql_str(machine)}, {packs}, "
                f"{sql_str(GO_LIVE_TS)}, NULL, "
                f"'Initial machine load (Go-Live)', {sql_str(CREATED_BY)})"
            )
        out.append(",\n".join(rows) + ";")
    else:
        out.append("-- (no machine stock - VMS sync จะเติมให้เอง)")
    out.append("")

    # ── 4) machine_stock (slot snapshot) ─────────────────────────
    if machine_slots:
        out.append("-- 4) machine_stock: snapshot ของแต่ละ slot")
        out.append("INSERT INTO machine_stock (machine_id, kiosk_record_id, slot_number,")
        out.append("                            sku_id, remain, max_capacity, status, synced_at)")
        out.append("VALUES")
        rows = []
        for s in machine_slots:
            rows.append(
                f"  ({sql_str(s['machine_id'])}, 0, {sql_str(s['slot_number'])}, "
                f"{sql_str(s['sku_id'])}, {s['remain']}, {s['max_capacity']}, "
                f"'active', {sql_str(GO_LIVE_TS)})"
            )
        out.append(",\n".join(rows) + ";")
        out.append("-- ⚠ kiosk_record_id ตั้งเป็น 0 ชั่วคราว · VMS sync ครั้งถัดไปจะอัปเดตให้ตรง")
    else:
        out.append("-- 4) machine_stock: ไม่ seed (VMS sync จะเติมตอน sync ครั้งถัดไป)")
    out.append("")

    # ── 5) avg_cost ──────────────────────────────────────────────
    out.append("-- 5) อัปเดต skus.avg_cost = unit_cost ของ stock_in (single lot)")
    rows = []
    for sku in sorted(all_skus):
        total = main.get(sku, 0) + user_sum.get(sku, 0) + machine_sum.get(sku, 0)
        if total <= 0:
            continue
        c = cost.get(sku, 0.0)
        rows.append(f"UPDATE skus SET avg_cost = {c:.2f} WHERE sku_id = {sql_str(sku)};")
    if rows:
        out.extend(rows)
    out.append("")

    # ── 6) Verify ────────────────────────────────────────────────
    out.append("-- 6) Verify: balance ของแต่ละ SKU ต้องตรง")
    out.append("DO $$")
    out.append("DECLARE")
    out.append("  v_diff INTEGER;")
    out.append("BEGIN")
    out.append("  SELECT COUNT(*) INTO v_diff")
    out.append("  FROM v_stock_balance vsb")
    out.append("  WHERE vsb.balance < 0;")
    out.append("  IF v_diff > 0 THEN")
    out.append("    RAISE EXCEPTION 'Seed failed: % SKU มี balance ติดลบ', v_diff;")
    out.append("  END IF;")
    out.append("  RAISE NOTICE 'Seed OK';")
    out.append("END $$;")
    out.append("")
    out.append("-- ถ้าทุกอย่างถูกต้อง → COMMIT")
    out.append("COMMIT;")
    out.append("-- ROLLBACK;")
    out.append("")

    # ── 7) summary ───────────────────────────────────────────────
    out.append("-- =============================================================")
    out.append("-- Summary จากไฟล์ Excel:")
    out.append(f"--   SKUs ที่มีของ:    {len([s for s in all_skus if (main.get(s,0)+user_sum.get(s,0)+machine_sum.get(s,0)) > 0])}")
    out.append(f"--   stock_in rows:     {len([s for s in all_skus if (main.get(s,0)+user_sum.get(s,0)+machine_sum.get(s,0)) > 0])}")
    out.append(f"--   stock_transfers:   {len(user_packs)}")
    out.append(f"--   stock_out:         {len(machine_packs)}")
    out.append(f"--   machine_stock:     {len(machine_slots)}")
    total_packs = sum(main.values()) + sum(user_packs.values()) + sum(machine_packs.values())
    out.append(f"--   รวมซองทั้งหมด:    {total_packs}")
    out.append("-- =============================================================")

    return "\n".join(out)


def main():
    in_path  = sys.argv[1] if len(sys.argv) > 1 else "golive_template.xlsx"
    out_path = "03_seed_generated.sql"

    wb = load_workbook(in_path, data_only=False)
    main_packs, cost, e1   = parse_main(wb["Main_Stock"])
    user_packs, e2         = parse_user(wb["User_Stock"])
    machine_packs, slots, e3 = parse_machine(wb["Machine_Stock"])

    errors = e1 + e2 + e3
    if errors:
        print("ERRORS — fix Excel ก่อนนำไป seed:", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        sys.exit(1)

    sql = gen_sql(main_packs, cost, user_packs, machine_packs, slots)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(sql)
    print(f"OK wrote {out_path}")
    user_sum    = sum(user_packs.values())
    machine_sum = sum(machine_packs.values())
    print(f"  - SKUs: {len(main_packs)}")
    print(f"  - Main packs: {sum(main_packs.values())}")
    print(f"  - User packs: {user_sum}")
    print(f"  - Machine packs: {machine_sum}")
    print(f"  - Total: {sum(main_packs.values()) + user_sum + machine_sum}")


if __name__ == "__main__":
    main()
